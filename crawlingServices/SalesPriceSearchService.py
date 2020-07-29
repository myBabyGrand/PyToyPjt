from selenium import webdriver
from selenium.webdriver.chrome.options import Options

from bs4 import BeautifulSoup
import time
from datetime import datetime
from openpyxl import Workbook
import os
import shutil

#workspace
workspacePath = ''
#웹드라이버 경로
webDriverPath = workspacePath+'/webDriver'
#관련 파일들의 default 경로
prsnDir = workspacePath+'/crawlingServices'
#key word
keyWordFile = prsnDir+'/salesPriceSearch/Keyword.txt'
keyWordFileBackupDir = prsnDir+'/salesPriceSearch/backup'

options = Options()
options.headless = True #창을 띄우지 않음

browser = webdriver.Chrome(executable_path=webDriverPath+'/chromedriver.exe', options=options)

#file util
def backUpAllFile(oldPath, newPath, timeStampAdd, isCopy):
    NOW = datetime.now()
    if os.path.exists(oldPath):
        print(checkNCreateDir(newPath))
        i = 0
        for eachFile in os.listdir(oldPath):
            print(os.path.isfile(oldPath+'/'+eachFile))
            if os.path.isfile(oldPath+'/'+eachFile) :
                newFileNm = str(eachFile)
                if timeStampAdd =='Y':
                    newFileNm = getSysdateAsStr('YYYYMMDDHH24MISS')+'_'+eachFile
                if isCopy :
                    shutil.copy(oldPath + '/' + eachFile, newPath + '/' + newFileNm) #copy
                else :
                    shutil.move(oldPath + '/' + eachFile, newPath + '/' + newFileNm) #move
                i += 1
        if i==0:
            return 'Directory exist but empty'
        else:
            return 'backup All'
    else:
        return 'Old Directory is not Exist'

def checkNCreateDir(chkPath):
    if os.path.exists(chkPath):
        return 'already exist'
    else:
        os.mkdir(chkPath)
        return 'Directory has Created! : ' + str(chkPath)

def getSysdateAsStr(type):
    NOW = datetime.now()
    if(type == 'YYYYMMDDHH24MISS'):
        return str(datetime.now())[:19].replace(' ','').replace('-','').replace(':','')
    else:
        return str(NOW)

#텍스트 파일에서 읽어오기
if not os.path.exists(keyWordFile) :
    print('keyword file does not exists : ' + keyWordFile)
    os._exit(1)

keyWordList = []

f = open(keyWordFile, 'r', encoding='utf-8')

keyWordList = f.readlines()
if len(keyWordList) > 0 :
    # 파일 읽기, 분기 시작

    # 엑셀파일
    excelFileNm = 'SalesPriceSearchResult.xlsx'
    wb = Workbook()

    danawaSearch = 'http://search.danawa.com/dsearch.php?k1='

    for keyword in keyWordList:
        print('keyword is ' + keyword)

        # for-loop : each keyword
        # keyword = 'gtx1660ti'
        url = danawaSearch + keyword
        browser.get(url)
        res = browser.page_source
        time.sleep(5)

        soup = BeautifulSoup(res, 'html.parser')
        searchResult = soup.find_all('li', {'class': 'prod_item width_change'})
        print(len(searchResult))

        ws = wb.create_sheet(keyword)

        # 헤더
        ws['A1'] = '품명'
        ws['B1'] = '가격'
        ws['C1'] = 'URL'
        ws['D1'] = 'Spec.'

        line = 1

        for item in searchResult:
            line = line + 1
            strItmNm = str(item.find('p', {'class': 'prod_name'}).text).rstrip()
            # print(strItmNm)
            strURL = str(item.find('p', {'class': 'prod_name'}).find('a').get('href')).rstrip()
            # print(strURL)
            strPrc = str(item.find('p', {'class': 'price_sect'}).text).strip()
            wonIdx = strPrc.find('원')
            # print(wonIdx)
            strPrc = strPrc[:wonIdx]
            # print(strPrc)
            spec = item.find('div', {'class': 'spec_list'})
            spec = spec.findAll('a')
            strSpec = ''
            for s in spec:
                val = s.get_text()
                strSpec = strSpec + str(val).strip() + ' '

            strSpec = str(strSpec).strip()
            # print(strSpec)
            ws.append([strItmNm, strPrc, strURL, strSpec])

    # 기존 파일 백업 - 이미 생성된 엑셀파일
    checkNCreateDir(prsnDir + '/output')
    checkNCreateDir(prsnDir + '/output/backup')
    print(backUpAllFile(prsnDir+'/output', prsnDir+'/output/backup', 'Y', False))

    # 파일 쓰기
    wb.save(os.path.join(prsnDir,'output', excelFileNm))

    f.close()
    # 기존 파일 백업 - keyword파일
    checkNCreateDir(keyWordFileBackupDir)
    print(backUpAllFile(prsnDir+'/salesPriceSearch/', keyWordFileBackupDir, 'Y', True))

else :
    print('keyword file is empty')
    f.close()
    os._exit(1)





